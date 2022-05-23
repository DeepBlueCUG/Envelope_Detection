# Postprocessing functions, plot graphs and write the results to excel files
import matplotlib.pyplot as plt
import openpyxl
import time
import os

def plot_data(data):

    data = data.tolist()

    plt.clf()
    for i in range(len(data)):
        plt.plot(data[i], ls="-", lw=2)


    plt.show()
    # plt.legend()
    # plt.xticks(x, names, rotation=1)

    plt.margins(0)
    plt.subplots_adjust(bottom=0.10)
    plt.xlabel('Time')
    plt.ylabel("NDVI")
    #plt.yticks([0.750, 0.800, 0.850])
    plt.title("Reconstrcution Compration")
    #plt.savefig('D:\\f1.jpg', dpi=900)

    #time.sleep(1.5)
    # plt.pause(2)
    plt.close()

def write_data_into_excel(data, sheet_name, output_path, output_file_name):

    output_length = len(data)

    new_excel = True

    for root, folders, files in os.walk(output_path):

        for item in files:
            if item == output_file_name:
                new_excel = False
                wb = openpyxl.load_workbook(output_path + output_file_name)
                sheet_names = wb.sheetnames
                new_sheet = True

                for check in sheet_names:
                    if check == sheet_name:
                        new_sheet = False
                        sheet = wb.get_sheet_by_name(check)
                        for i in range(output_length):
                            current_timeseries = data[i]
                            for j in range(len(current_timeseries)):
                                sheet.cell(j + 1, i + 1, current_timeseries[j])
                        wb.save(output_path + output_file_name)
                        wb.close()

                if new_sheet:
                    sheet = wb.create_sheet(sheet_name)
                    for i in range(output_length):
                        current_timeseries = data[i]
                        for j in range(len(current_timeseries)):
                            sheet.cell(j + 1, i + 1, current_timeseries[j])
                    wb.save(output_path + output_file_name)
                    wb.close()

    if new_excel:

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = sheet_name
        for i in range(output_length):
            current_timeseries = data[i]
            for j in range(len(current_timeseries)):
                sheet.cell(j + 1, i + 1, current_timeseries[j])
        book.save(output_path + output_file_name)
        book.close()
